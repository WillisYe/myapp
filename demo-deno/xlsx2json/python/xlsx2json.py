#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
XLSX转JSON脚本 - Python版本
用法: python xlsx2json.py <输入文件路径> [输出文件路径]

将XLSX文件的第一个工作表转换为JSON格式
A列作为key，B列作为value
"""

import sys
import json
import os
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("❌ 缺少依赖包 openpyxl")
    print("请运行: pip install openpyxl")
    sys.exit(1)


def show_usage():
    print("用法:")
    print("  python xlsx2json.py <输入文件路径> [输出文件路径]")
    print("")
    print("参数:")
    print("  输入文件路径    - 要转换的XLSX文件路径")
    print("  输出文件路径    - 生成的JSON文件路径（可选，默认与输入文件同名但扩展名为.json）")
    print("")
    print("示例:")
    print("  python xlsx2json.py data.xlsx")
    print("  python xlsx2json.py data.xlsx output.json")
    print("")
    print("安装依赖:")
    print("  pip install -r requirements.txt")


def validate_file(file_path):
    """验证输入文件"""
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"文件不存在: {file_path}")

    if not os.path.isfile(file_path):
        raise ValueError(f"{file_path} 不是一个文件")

    ext = Path(file_path).suffix.lower()
    if ext not in ['.xlsx', '.xls']:
        raise ValueError(f"{file_path} 不是Excel文件格式（需要.xlsx或.xls）")

    return True


def generate_output_path(input_path):
    """生成输出文件路径"""
    path = Path(input_path)
    return str(path.with_suffix('.json'))


def convert_xlsx_to_json(input_path, output_path):
    """转换XLSX文件为JSON"""
    try:
        print(f"正在读取文件: {input_path}")

        # 读取Excel文件
        workbook = openpyxl.load_workbook(input_path, data_only=True)

        # 获取第一个工作表
        sheet_names = workbook.sheetnames
        if not sheet_names:
            raise ValueError("Excel文件中没有找到工作表")

        first_sheet_name = sheet_names[0]
        worksheet = workbook[first_sheet_name]

        print(f"正在处理工作表: {first_sheet_name}")

        # 转换为key-value对象
        result = {}
        processed_rows = 0

        for row_num, row in enumerate(worksheet.iter_rows(values_only=True), 1):
            # 跳过空行
            if not row or all(cell is None or cell == "" for cell in row):
                continue

            # 确保至少有两列数据
            if len(row) < 2:
                continue

            key = row[0]  # A列作为key
            value = row[1]  # B列作为value

            # 跳过key为空的行
            if key is None or key == "":
                continue

            # 转换为字符串
            key_str = str(key).strip()
            value_str = str(value).strip() if value is not None else ""

            result[key_str] = value_str
            processed_rows += 1

        print(f"已处理 {processed_rows} 行数据")

        if processed_rows == 0:
            print("警告: 没有找到有效的数据行（A列为空）")

        # 写入JSON文件
        print(f"正在写入文件: {output_path}")
        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(result, f, ensure_ascii=False, indent=2)

        print("✅ 转换完成！")
        print(f"   输入文件: {input_path}")
        print(f"   输出文件: {output_path}")
        print(f"   转换行数: {processed_rows}")

    except Exception as error:
        print(f"❌ 转换失败: {error}")
        sys.exit(1)


def main():
    """主函数"""
    args = sys.argv[1:]

    if len(args) == 0 or '-h' in args or '--help' in args:
        show_usage()
        return

    if len(args) < 1 or len(args) > 2:
        print("❌ 参数数量错误")
        show_usage()
        sys.exit(1)

    input_path = args[0]
    output_path = args[1] if len(args) > 1 else generate_output_path(input_path)

    try:
        # 验证输入文件
        validate_file(input_path)

        # 执行转换
        convert_xlsx_to_json(input_path, output_path)

    except Exception as error:
        print(f"❌ {error}")
        sys.exit(1)


if __name__ == "__main__":
    main()